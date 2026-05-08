import os
import unicodedata
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from employees.models import Department, Employee
from accounts.models import User

DEPT_FILE = os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'sample_input', 'department.xlsx'
)

DEFAULT_PASSWORD = '123456'


def _to_ascii_slug(name: str) -> str:
    """Convert Vietnamese name to ASCII slug (last word / given name)."""
    parts = name.strip().split()
    given_name = parts[-1] if parts else name
    normalized = unicodedata.normalize('NFD', given_name)
    ascii_only = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
    return ascii_only.lower().replace(' ', '')


def _build_email_map(rows):
    """Return dict: employee_code -> email, handling duplicate given names."""
    slug_count = {}
    for _, full_name, _, _ in rows:
        slug = _to_ascii_slug(full_name)
        slug_count[slug] = slug_count.get(slug, 0) + 1

    slug_seen = {}
    email_map = {}
    for code, full_name, _, _ in rows:
        slug = _to_ascii_slug(full_name)
        if slug_count[slug] == 1:
            email = f'{slug}@gmail.com'
        else:
            slug_seen[slug] = slug_seen.get(slug, 0) + 1
            email = f'{slug}{slug_seen[slug]}@gmail.com'
        email_map[code] = email
    return email_map


class Command(BaseCommand):
    help = 'Seed employees from sample_input/department.xlsx'

    def handle(self, *args, **options):
        path = os.path.abspath(DEPT_FILE)
        if not os.path.exists(path):
            self.stderr.write(f'File not found: {path}')
            return

        wb = load_workbook(path, data_only=True)
        ws = wb.active

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            code = row[0]
            full_name = row[1]
            position = row[2]
            dept_name = row[3]
            if not code or not full_name or not dept_name:
                continue
            rows.append((str(code).strip(), str(full_name).strip(),
                         str(position).strip() if position else '', str(dept_name).strip()))

        email_map = _build_email_map(rows)
        created = 0
        skipped = 0

        with transaction.atomic():
            for code, full_name, position, dept_name in rows:
                if Employee.objects.filter(code=code).exists():
                    skipped += 1
                    continue

                dept, _ = Department.objects.get_or_create(name=dept_name)
                email = email_map[code]

                # Ensure email is unique (edge case: same slug across different rows)
                base_email = email
                suffix = 0
                while User.objects.filter(email=email).exists():
                    suffix += 1
                    name_part = base_email.split('@')[0]
                    email = f'{name_part}x{suffix}@gmail.com'

                user = User.objects.create_user(
                    email=email,
                    password=DEFAULT_PASSWORD,
                    role=User.Role.EMPLOYEE,
                )
                Employee.objects.create(
                    user=user,
                    code=code,
                    full_name=full_name,
                    department=dept,
                    position=position,
                )
                created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Seeded employees: {created} created, {skipped} already existed.'
        ))
        self.stdout.write(f'Default password for all: {DEFAULT_PASSWORD}')
