import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from employees.models import Department

DEPT_FILE = os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'sample_input', 'department.xlsx'
)


class Command(BaseCommand):
    help = 'Seed departments from sample_input/department.xlsx'

    def handle(self, *args, **options):
        path = os.path.abspath(DEPT_FILE)
        if not os.path.exists(path):
            self.stderr.write(f'File not found: {path}')
            return

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        names = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            dept_name = row[3]
            if dept_name:
                names.add(str(dept_name).strip())

        created = 0
        for name in sorted(names):
            _, was_created = Department.objects.get_or_create(name=name)
            if was_created:
                created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Seeded departments: {created} created, {len(names) - created} already existed.'
        ))
