from dataclasses import dataclass
from datetime import date, time, datetime
from typing import Optional
from openpyxl import load_workbook

# Expected header columns in order (index: name)
REQUIRED_COLUMNS = {
    0: 'Mã nhân viên',
    2: 'Họ và tên',
    3: 'Ngày công',
    5: 'Ca làm việc đăng ký',
    10: 'Giờ vào Máy chấm công',
    11: 'Giờ ra Máy chấm công',
}


@dataclass
class AttendanceRow:
    employee_code: str
    employee_name: str
    date: date
    check_in: Optional[time]
    check_out: Optional[time]
    shift_code: str


class InvalidAttendanceFile(Exception):
    pass


def _parse_time(value) -> Optional[time]:
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                return datetime.strptime(value, fmt).time()
            except ValueError:
                pass
    return None


def _parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _validate_header(header_row):
    """Raise InvalidAttendanceFile if required columns are missing."""
    missing = []
    for idx, expected_name in REQUIRED_COLUMNS.items():
        actual = header_row[idx] if idx < len(header_row) else None
        if actual is None or str(actual).strip() != expected_name:
            missing.append(f'Cột {idx + 1}: "{expected_name}" (thực tế: "{actual}")')
    if missing:
        raise InvalidAttendanceFile(
            'File không đúng định dạng. Các cột bị thiếu hoặc sai:\n' + '\n'.join(missing)
        )


def parse_attendance_excel(file_path: str) -> list:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            _validate_header(row)
            continue
        if not row[0]:
            continue
        emp_code = str(row[0]).strip()
        if not emp_code:
            continue
        parsed_date = _parse_date(row[3])
        if not parsed_date:
            continue
        rows.append(AttendanceRow(
            employee_code=emp_code,
            employee_name=str(row[2]).strip() if row[2] else '',
            date=parsed_date,
            check_in=_parse_time(row[10]),
            check_out=_parse_time(row[11]),
            shift_code=str(row[5]).strip() if row[5] else '',
        ))
    return rows
