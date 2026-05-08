import os
from datetime import time
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from attendance.models import Shift

SHIFT_FILE = os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'sample_input', 'shift.xlsx'
)

# Columns: STT, Ký hiệu ca, Giờ vào, Giờ ra, Tổng thời gian, Ngày công,
#          Ngày phép, Giờ công, Giờ phép, Ghi chú, Nghỉ trưa bắt đầu, Nghỉ trưa kết thúc


def _to_time(val):
    if val is None or val == 'NA':
        return None
    if isinstance(val, time):
        return val
    return None


def _to_decimal(val):
    if val is None:
        return 0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0


class Command(BaseCommand):
    help = 'Seed shifts from sample_input/shift.xlsx'

    def handle(self, *args, **options):
        path = os.path.abspath(SHIFT_FILE)
        if not os.path.exists(path):
            self.stderr.write(f'File not found: {path}')
            return

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        created = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            code = row[1]
            if not code:
                continue
            code = str(code).strip()
            if not code:
                continue

            defaults = {
                'check_in': _to_time(row[2]),
                'check_out': _to_time(row[3]),
                'total_hours': _to_decimal(row[4]),
                'workday_value': _to_decimal(row[5]),
                'leave_day_value': _to_decimal(row[6]),
                'work_hours': _to_decimal(row[7]),
                'leave_hours': _to_decimal(row[8]),
                'note': str(row[9]).strip() if row[9] else '',
                'break_start': _to_time(row[10]),
                'break_end': _to_time(row[11]),
            }
            _, was_created = Shift.objects.get_or_create(code=code, defaults=defaults)
            if was_created:
                created += 1
            else:
                skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f'Seeded shifts: {created} created, {skipped} already existed.'
        ))
