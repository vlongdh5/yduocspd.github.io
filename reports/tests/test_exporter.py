import pytest
from reports.exporter import export_calculation_excel
from reports.models import AttendanceCalculation
from employees.models import Employee, Department, LeaveBalance
from accounts.models import User
import tempfile, os
from openpyxl import load_workbook

@pytest.fixture
def setup(db):
    user = User.objects.create_user(email='emp@example.com', password='pass')
    dept = Department.objects.create(name='KD')
    emp = Employee.objects.create(user=user, code='NV001', full_name='Nguyen Van A', department=dept)
    hr = User.objects.create_user(email='hr@example.com', password='pass', role=User.Role.HR)
    LeaveBalance.objects.create(employee=emp, year=2026, total_days=12, used_hours=16)
    AttendanceCalculation.objects.create(
        employee=emp, month='2026-05', actual_workdays=20, leave_days_used=2, calculated_by=hr
    )
    return {'emp': emp}

@pytest.mark.django_db
def test_export_creates_file(setup):
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    export_calculation_excel('2026-05', tmp.name)
    wb = load_workbook(tmp.name)
    os.unlink(tmp.name)
    assert wb is not None
    ws = wb.active
    assert ws.max_row >= 2

@pytest.mark.django_db
def test_export_contains_employee_data(setup):
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    export_calculation_excel('2026-05', tmp.name)
    wb = load_workbook(tmp.name)
    os.unlink(tmp.name)
    ws = wb.active
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    flat = [str(v) for row in data for v in row if v is not None]
    assert 'NV001' in flat
    assert 'Nguyen Van A' in flat
