from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reports.models import AttendanceCalculation
from employees.models import LeaveBalance

_ERROR_LABELS = {
    'LATE': 'Đi muộn',
    'MISSING_IN': 'Thiếu giờ vào',
    'EARLY_LEAVE': 'Về sớm',
    'MISSING_OUT': 'Thiếu giờ ra',
    'ABSENT': 'Vắng mặt',
}

_STATUS_LABELS = {
    'pending': 'Đang chờ',
    'approved': 'Đã duyệt',
    'rejected': 'Từ chối',
}


def _error_position(error_types):
    es = set(error_types or [])
    if 'ABSENT' in es:
        return 'Cả ngày'
    has_ci = bool(es & {'LATE', 'MISSING_IN'})
    has_co = bool(es & {'EARLY_LEAVE', 'MISSING_OUT'})
    if has_ci and has_co:
        return 'Đầu ca & Cuối ca'
    if has_ci:
        return 'Đầu ca'
    if has_co:
        return 'Cuối ca'
    return '-'


def _build_detail_rows(month: str):
    from attendance.models import AttendanceRecord, Shift
    from explanations.models import Explanation
    from reports.calculator import compute_record_hours, _is_qcc_record

    records = (
        AttendanceRecord.objects.filter(upload__month=month)
        .select_related(
            'employee__department',
            'explanation__ci_reason',
            'explanation__co_reason',
        )
        .order_by('employee__code', 'date')
    )
    shift_map = {s.code: s for s in Shift.objects.filter(is_active=True)}
    qcc_by_emp = {}
    rows = []

    for record in records:
        emp = record.employee
        exp = getattr(record, 'explanation', None)
        shift = shift_map.get(record.shift_code) if record.shift_code else None

        qcc_count = qcc_by_emp.get(emp.code, 0)
        if _is_qcc_record(exp):
            qcc_count += 1
            qcc_by_emp[emp.code] = qcc_count

        w, l = compute_record_hours(record, exp, shift, qcc_count)

        error_types = record.error_types or []
        error_label = ', '.join(_ERROR_LABELS.get(e, e) for e in error_types) if error_types else '-'

        ci_reason = exp.ci_reason.name if exp and exp.ci_reason else '-'
        ci_status_raw = exp.ci_status if exp and exp.ci_status else None
        ci_status = _STATUS_LABELS.get(ci_status_raw, '-') if ci_status_raw else '-'
        co_reason = exp.co_reason.name if exp and exp.co_reason else '-'
        co_status_raw = exp.co_status if exp and exp.co_status else None
        co_status = _STATUS_LABELS.get(co_status_raw, '-') if co_status_raw else '-'

        rows.append({
            'emp_code': emp.code,
            'emp_name': emp.full_name,
            'dept': emp.department.name if emp.department_id else '',
            'date': record.date,
            'shift_code': record.shift_code or '-',
            'error_label': error_label,
            'minutes_late': record.minutes_late if record.minutes_late is not None else '-',
            'minutes_early': record.minutes_early if record.minutes_early is not None else '-',
            'error_position': _error_position(error_types),
            'ci_reason': ci_reason,
            'ci_status': ci_status,
            'ci_status_raw': ci_status_raw,
            'co_reason': co_reason,
            'co_status': co_status,
            'co_status_raw': co_status_raw,
            'work_hours': w,
            'leave_hours': l,
            'has_error': bool(error_types),
            'all_approved': (
                (not record.has_ci_issue or ci_status_raw == Explanation.Status.APPROVED) and
                (not record.has_co_issue or co_status_raw == Explanation.Status.APPROVED)
            ) if bool(error_types) else True,
        })

    return rows


def _apply_border(ws, row_idx, num_cols, border, alignments=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = border
        if alignments and col in alignments:
            cell.alignment = alignments[col]


def export_calculation_excel(month: str, output_path: str):
    calcs = AttendanceCalculation.objects.filter(month=month).select_related(
        'employee__department'
    ).order_by('employee__department__name', 'employee__code')

    wb = Workbook()

    # ── Sheet 1: Tổng hợp ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f'Tổng hợp {month}'

    title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    header_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center')

    ws1.merge_cells('A1:I1')
    ws1['A1'] = f'BẢNG TỔNG HỢP CÔNG THÁNG {month}'
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30
    ws1.append([])

    headers1 = ['STT', 'Mã NV', 'Họ tên', 'Phòng ban', 'Giờ công', 'Giờ phép', 'Ngày công (÷8)', 'Phép còn lại', 'Ghi chú']
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for i, calc in enumerate(calcs, 1):
        emp = calc.employee
        try:
            lb = LeaveBalance.objects.get(employee=emp, year=int(month[:4]))
            remaining = float(lb.remaining_days)
        except LeaveBalance.DoesNotExist:
            remaining = '-'

        row_data = [
            i, emp.code, emp.full_name,
            emp.department.name if emp.department_id else '',
            float(calc.work_hours), float(calc.leave_hours),
            float(calc.actual_workdays), remaining, '',
        ]
        ws1.append(row_data)
        row_idx = ws1.max_row
        center_cols = {1, 5, 6, 7, 8}
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col)
            cell.border = border
            if col in center_cols:
                cell.alignment = center

    for col, width in enumerate([6, 10, 25, 20, 12, 12, 14, 14, 20], 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # ── Sheet 2: Chi tiết từng ngày ────────────────────────────────────────
    ws2 = wb.create_sheet(title=f'Chi tiết {month}')

    ws2.merge_cells('A1:O1')
    ws2['A1'] = f'CHI TIẾT CHẤM CÔNG TỪNG NGÀY THÁNG {month}'
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30
    ws2.append([])

    headers2 = [
        'STT', 'Mã NV', 'Họ tên', 'Phòng ban', 'Ngày', 'Ca',
        'Loại lỗi', 'Phút vào muộn', 'Phút ra sớm', 'Vị trí lỗi',
        'Lý do giải trình vào', 'Phê duyệt vào',
        'Lý do giải trình ra', 'Phê duyệt ra',
        'Giờ công', 'Giờ phép',
    ]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws2.row_dimensions[3].height = 30

    # Fill colours for detail rows
    ok_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    approved_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # light green
    pending_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')   # light yellow
    rejected_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # light red

    detail_rows = _build_detail_rows(month)
    center_cols2 = {1, 2, 5, 6, 8, 9, 10, 12, 14, 15, 16}

    for i, r in enumerate(detail_rows, 1):
        row_data = [
            i, r['emp_code'], r['emp_name'], r['dept'],
            r['date'].strftime('%d/%m/%Y'), r['shift_code'],
            r['error_label'],
            r['minutes_late'], r['minutes_early'], r['error_position'],
            r['ci_reason'], r['ci_status'],
            r['co_reason'], r['co_status'],
            r['work_hours'], r['leave_hours'],
        ]
        ws2.append(row_data)
        row_idx = ws2.max_row

        # Row colour based on error/approval state
        if not r['has_error']:
            row_fill = ok_fill
        elif r['all_approved']:
            row_fill = approved_fill
        elif r['ci_status_raw'] == Explanation.Status.REJECTED or r['co_status_raw'] == Explanation.Status.REJECTED:
            row_fill = rejected_fill
        else:
            row_fill = pending_fill

        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col)
            cell.border = border
            cell.fill = row_fill
            if col in center_cols2:
                cell.alignment = center

    col_widths2 = [6, 10, 25, 18, 12, 14, 20, 14, 14, 18, 28, 14, 28, 14, 12, 12]
    for col, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)
